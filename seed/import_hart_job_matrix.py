#!/usr/bin/env python3
"""Import job steps and pickup criteria from hart_job_criteria_matrix.xlsx into config."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

SEED_DIR = Path(__file__).resolve().parent
REPO_ROOT = SEED_DIR.parent
DEFAULT_CONFIG = SEED_DIR / "hart_seed_config.json"
DEFAULT_MATRIX = SEED_DIR / "hart_job_criteria_matrix.xlsx"

JOB_DESCRIPTIONS: dict[str, str] = {
    "D749": (
        "CSX Neville Island Switcher.\n"
        "Works South Yard and CSX Demmler interchange.\n"
        "- Demmler Yard: pick up inbound and offline for Scully, Shenango, Neville Island, and Demmler; set out outbound\n"
        "- South Yard: set out inbound; pick up for Demmler"
    ),
    "NVL": (
        "POHC Neville Local.\n"
        "Works Scully interchange, Neville Island industries, and South Yard.\n"
        "- Scully Yard: pick up and set out interchange traffic for all destinations\n"
        "- Neville Island: industry spotting and pulls for Scully, Shenango, island, and Demmler\n"
        "- South Yard: pick up blocks for Scully and island; set out staging"
    ),
    "YM1": (
        "South Yard yardmaster (YM1) — inter-island switching on Neville Island.\n"
        "Retrieve and stage cars across North, West, and East satellite yards for island industries.\n"
        "Sort inbound traffic and build blocks for Scully, Shenango, South Yard, and Demmler (CSX D749).\n"
        "South Yard staging completes blocks for POHC NVL (via Scully).\n"
        "D749 and NVL handle interchange and industry spotting; YM1 works satellite yards only."
    ),
    "CK1": (
        "Coke transfer — optional yard move.\n"
        "Move coke loads between Shenango Coke Works, North Yard, and South Yard for weighing and classification when authorized.\n"
        "North Yard and Shenango export coke to Demmler Offline and Scully Offline; South Yard picks for North Yard and Shenango only.\n"
        "Run only when it will not interfere with NVL or passenger movements."
    ),
    "STG-SCULLY": (
        "Scully Yard staging — shuffle to Scully Offline, stage blocks, set out at each step.\n"
        "Step 10 picks at Scully Yard; step 11 set out at Scully Offline; steps 12–50 pick and set out at Scully Offline; step 60 set out at Scully Yard."
    ),
    "STG-DEMMLER": (
        "Demmler Yard staging — shuffle to Demmler Offline, stage blocks, set out at each step.\n"
        "Step 10 picks at Demmler Yard; step 11 set out at Demmler Offline; steps 12–50 pick and set out at Demmler Offline; step 60 set out at Demmler Yard."
    ),
}


def parse_station_id(label: str) -> int:
    match = re.match(r"(\d+)", str(label).strip())
    if not match:
        raise ValueError(f"Cannot parse station id from label: {label!r}")
    return int(match.group(1))


def load_config(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def save_config(path: Path, config: dict) -> None:
    with path.open("w", encoding="utf-8") as handle:
        json.dump(config, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def station_maps(stations: list[dict]) -> tuple[dict[int, str], dict[int, int]]:
    names = {int(s["id"]): s["name"] for s in stations}
    sort_seq = {int(s["id"]): int(s.get("sort_seq", 999)) for s in stations}
    return names, sort_seq


def read_job_matrix(ws) -> dict[int, dict[int, str]]:
    header_row = None
    for row_idx in range(1, 25):
        value = ws.cell(row_idx, 1).value
        if value and "At station" in str(value):
            header_row = row_idx
            break
    if header_row is None:
        return {}

    dest_ids: list[int] = []
    for col_idx in range(2, 50):
        header = ws.cell(header_row, col_idx).value
        if header is None:
            break
        dest_ids.append(parse_station_id(str(header)))

    matrix: dict[int, dict[int, str]] = {}
    for row_idx in range(header_row + 1, header_row + 50):
        row_label = ws.cell(row_idx, 1).value
        if row_label is None:
            break
        work_id = parse_station_id(str(row_label))
        row_cells: dict[int, str] = {}
        for col_offset, dest_id in enumerate(dest_ids, start=2):
            raw = ws.cell(row_idx, col_offset).value
            if raw is None:
                continue
            value = str(raw).strip().upper()
            if value in {"P", "S", "SP"}:
                row_cells[dest_id] = value
        if row_cells:
            matrix[work_id] = row_cells
    return matrix


def build_pickup_remark(work_id: int, dest_id: int, names: dict[int, str]) -> str:
    return f"{names[work_id]} — Pick up for {names[dest_id]}"


def build_pickup_setout_remark(work_id: int, dest_id: int, names: dict[int, str]) -> str:
    return f"{names[work_id]} — Pick up for {names[dest_id]}; Set out at {names[work_id]}"


def build_setout_remark(work_id: int, names: dict[int, str]) -> str:
    return f"{names[work_id]} — Set out at {names[work_id]}"


def matrix_to_job_rules(
    job_name: str,
    matrix: dict[int, dict[int, str]],
    names: dict[int, str],
    sort_seq: dict[int, int],
) -> tuple[list[dict], list[dict]]:
    """Expand matrix to one step per pickup destination (single pu_criteria row each)."""
    steps: list[dict] = []
    criteria: list[dict] = []
    step_number = 10
    active_work_ids = sorted(matrix.keys(), key=lambda station_id: (sort_seq[station_id], station_id))

    for work_id in active_work_ids:
        cells = matrix[work_id]
        diagonal = cells.get(work_id, "")

        pickup_cells = sorted(
            ((dest_id, value) for dest_id, value in cells.items() if value in {"P", "SP"}),
            key=lambda item: (sort_seq[item[0]], item[0]),
        )

        for dest_id, value in pickup_cells:
            setout_on_step = value == "SP"
            remark = (
                build_pickup_setout_remark(work_id, dest_id, names)
                if setout_on_step
                else build_pickup_remark(work_id, dest_id, names)
            )
            steps.append(
                {
                    "step_number": step_number,
                    "station": work_id,
                    "pickup": True,
                    "setout": setout_on_step,
                    "remarks": remark,
                }
            )
            criteria.append(
                {
                    "job": job_name,
                    "step_nbr": step_number,
                    "car_status": "",
                    "dest_station_id": dest_id,
                }
            )
            step_number += 10

        if diagonal == "S":
            steps.append(
                {
                    "step_number": step_number,
                    "station": work_id,
                    "pickup": False,
                    "setout": True,
                    "remarks": build_setout_remark(work_id, names),
                }
            )
            step_number += 10

    return steps, criteria


def job_description(job_name: str) -> str:
    return JOB_DESCRIPTIONS.get(job_name, job_name)


def sql_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "''")


def bool_sql(flag: bool) -> str:
    return "T" if flag else "F"


def write_migration(config: dict, output_path: Path) -> None:
    lines = [
        "-- Apply matrix-derived job rules to live STS database.",
        "-- Generated from hart_job_criteria_matrix.xlsx via import_hart_job_matrix.py",
        "",
        "DELETE FROM pu_criteria;",
        "",
    ]
    for job in config["jobs"]:
        name = job["name"]
        lines.append(f"DELETE FROM `{name}`;")
        for step in job["steps"]:
            lines.append(
                "INSERT INTO `{name}` (step_number, station, pickup, setout, remarks) VALUES "
                "({step}, {station}, '{pickup}', '{setout}', '{remarks}');".format(
                    name=name,
                    step=int(step["step_number"]),
                    station=int(step["station"]),
                    pickup=bool_sql(bool(step["pickup"])),
                    setout=bool_sql(bool(step["setout"])),
                    remarks=sql_escape(step.get("remarks", "")),
                )
            )
        lines.append("")

    for job in config["jobs"]:
        desc = sql_escape(job.get("description", ""))
        lines.append(
            f"UPDATE jobs SET description = '{desc}' WHERE name = '{sql_escape(job['name'])}';"
        )
    lines.append("")

    criteria_id = 1
    for row in config.get("pickup_criteria", []):
        lines.append(
            "INSERT INTO pu_criteria (id, job_id, step_nbr, car_status, commodity_id, car_code_id, dest_station_id) "
            f"VALUES ({criteria_id}, '{sql_escape(row['job'])}', {int(row['step_nbr'])}, '', NULL, NULL, {int(row['dest_station_id'])});"
        )
        criteria_id += 1
    lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def import_matrix_to_config(
    config: dict,
    matrix_path: Path,
    job_sheet_names: list[str] | None = None,
) -> tuple[dict, list[dict]]:
    names, sort_seq = station_maps(config["stations"])
    workbook = load_workbook(matrix_path, data_only=True)

    jobs_by_name = {job["name"]: job for job in config["jobs"]}
    sheet_names = job_sheet_names or [name for name in workbook.sheetnames if name != "Legend"]

    all_criteria: list[dict] = []
    for sheet_name in sheet_names:
        if sheet_name not in jobs_by_name:
            raise KeyError(f"Sheet {sheet_name!r} has no matching job in config")
        matrix = read_job_matrix(workbook[sheet_name])
        steps, criteria = matrix_to_job_rules(sheet_name, matrix, names, sort_seq)
        job = jobs_by_name[sheet_name]
        job["steps"] = steps
        job["description"] = job_description(sheet_name)
        all_criteria.extend(criteria)

    config["pickup_criteria"] = all_criteria
    return config, all_criteria


def main(argv: list[str] | None = None) -> int:
    argv = argv or sys.argv[1:]
    config_path = Path(argv[0]) if argv else DEFAULT_CONFIG
    matrix_path = Path(argv[1]) if len(argv) > 1 else DEFAULT_MATRIX
    migration_path = Path(argv[2]) if len(argv) > 2 else REPO_ROOT / "migrations" / "matrix_import_migration.sql"

    config = load_config(config_path)
    config, criteria = import_matrix_to_config(config, matrix_path)
    save_config(config_path, config)
    write_migration(config, migration_path)

    print(f"Updated {config_path} from {matrix_path}")
    print(f"Wrote {migration_path}")
    for job in config["jobs"]:
        print(f"  {job['name']}: {len(job['steps'])} steps")
    print(f"  pickup_criteria: {len(criteria)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
