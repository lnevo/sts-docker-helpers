#!/usr/bin/env python3
"""Build an Excel station matrix per HART job for planning pickup/setout criteria.

Rows = station where the train works (job step location).
Columns = car destination station (pu_criteria dest_station_id).

Cells: P = pickup, S = set-out, SP = both.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SEED_DIR = Path(__file__).resolve().parent
DEFAULT_CONFIG = SEED_DIR / "hart_seed_config.json"
DEFAULT_OUTPUT = SEED_DIR / "hart_job_criteria_matrix.xlsx"

sys.path.insert(0, str(SEED_DIR))
from generate_hart_seed import SeedBuilder  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
LEGEND_FILL = PatternFill("solid", fgColor="FFF2CC")
P_FILL = PatternFill("solid", fgColor="C6EFCE")
S_FILL = PatternFill("solid", fgColor="FCE4D6")
SP_FILL = PatternFill("solid", fgColor="E2EFDA")


def load_config(path: Path) -> dict:
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def station_label(station: dict) -> str:
    return f"{station['id']} {station['name']}"


def build_pickup_criteria(config: dict) -> list[dict]:
    builder = SeedBuilder(config)
    builder.add_pickup_criteria()
    return builder.pu_criteria_rows


def merge_cell_value(current: str, addition: str) -> str:
    flags = set(current)
    flags.update(addition)
    if "P" in flags and "S" in flags:
        return "SP"
    if "P" in flags:
        return "P"
    if "S" in flags:
        return "S"
    return ""


def build_job_matrix(
    job: dict,
    criteria: list[dict],
    station_ids: list[int],
) -> dict[tuple[int, int], str]:
    """Return (work_station, dest_station) -> P|S|SP."""
    matrix: dict[tuple[int, int], str] = {}
    job_name = job["name"]
    steps_by_number = {step["step_number"]: step for step in job.get("steps", [])}

    def set_cell(work: int, dest: int, kind: str) -> None:
        key = (work, dest)
        matrix[key] = merge_cell_value(matrix.get(key, ""), kind)

    for row in criteria:
        if row["job"] != job_name:
            continue
        step = steps_by_number.get(int(row["step_nbr"]))
        if not step:
            continue
        work_station = int(step["station"])
        dest_station = int(row["dest_station_id"])
        set_cell(work_station, dest_station, "P")

    for step in job.get("steps", []):
        if not step.get("setout"):
            continue
        work_station = int(step["station"])
        # Set-out at the work station (cars spotted / released here).
        set_cell(work_station, work_station, "S")

    return matrix


def cell_fill(value: str) -> PatternFill | None:
    if value == "SP":
        return SP_FILL
    if value == "P":
        return P_FILL
    if value == "S":
        return S_FILL
    return None


def write_legend_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Legend", 0)
    sheet["A1"] = "HART job criteria matrix"
    sheet["A1"].font = Font(bold=True, size=14)
    lines = [
        ("", ""),
        ("Rows (At station)", "Station where the train works on a job step."),
        ("Columns (Car destination)", "Destination station matched by pickup criteria."),
        ("P", "Pick up cars at the row station destined for the column station."),
        ("S", "Set out cars at the row station (same row/column)."),
        ("SP", "Both pick up and set out on that station pair."),
        ("", ""),
        ("Notes", "Sequence numbers are intentionally omitted — map moves here first, then assign steps."),
        ("", "Blank cells = no criteria for that pair on this job."),
        ("", "After editing: python3 import_hart_job_matrix.py → updates config + migration SQL."),
        ("", "Then: python3 generate_hart_seed.py to refresh sts-backups/hart_seed."),
    ]
    for row_idx, (label, detail) in enumerate(lines, start=2):
        sheet.cell(row=row_idx, column=1, value=label)
        sheet.cell(row=row_idx, column=2, value=detail)
        if label in {"P", "S", "SP"}:
            fill = cell_fill(label)
            if fill:
                sheet.cell(row=row_idx, column=1).fill = fill
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 72


def write_job_sheet(
    workbook: Workbook,
    job: dict,
    stations: list[dict],
    station_ids: list[int],
    matrix: dict[tuple[int, int], str],
) -> None:
    name = job["name"]
    sheet = workbook.create_sheet(name)
    sheet["A1"] = f"{name} — station matrix"
    sheet["A1"].font = Font(bold=True, size=12)
    description = job.get("description", "").replace("\n", " ")
    sheet["A2"] = description[:240]
    sheet["A2"].alignment = Alignment(wrap_text=True)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(station_ids) + 2)

    header_row = 4
    sheet.cell(row=header_row, column=1, value="At station \\ Car destination")
    sheet.cell(row=header_row, column=1).font = Font(bold=True)
    sheet.cell(row=header_row, column=1).fill = HEADER_FILL

    for col_idx, station_id in enumerate(station_ids, start=2):
        station = next(s for s in stations if s["id"] == station_id)
        cell = sheet.cell(row=header_row, column=col_idx, value=station_label(station))
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_offset, work_id in enumerate(station_ids, start=1):
        row_idx = header_row + row_offset
        work_station = next(s for s in stations if s["id"] == work_id)
        label_cell = sheet.cell(row=row_idx, column=1, value=station_label(work_station))
        label_cell.font = Font(bold=True)
        label_cell.fill = HEADER_FILL

        for col_offset, dest_id in enumerate(station_ids, start=2):
            value = matrix.get((work_id, dest_id), "")
            cell = sheet.cell(row=row_idx, column=col_offset, value=value)
            cell.alignment = Alignment(horizontal="center")
            fill = cell_fill(value)
            if fill:
                cell.fill = fill

    sheet.freeze_panes = "B5"
    sheet.column_dimensions["A"].width = 26
    for col_idx in range(2, len(station_ids) + 2):
        sheet.column_dimensions[get_column_letter(col_idx)].width = 16
    sheet.row_dimensions[header_row].height = 36


def generate_workbook(config_path: Path, output_path: Path) -> None:
    config = load_config(config_path)
    stations = sorted(config["stations"], key=lambda s: (s.get("sort_seq", 999), s["id"]))
    station_ids = [int(s["id"]) for s in stations]
    criteria = build_pickup_criteria(config)

    workbook = Workbook()
    workbook.remove(workbook.active)
    write_legend_sheet(workbook)

    for job in config["jobs"]:
        matrix = build_job_matrix(job, criteria, station_ids)
        write_job_sheet(workbook, job, stations, station_ids, matrix)

    workbook.save(output_path)


def main(argv: list[str] | None = None) -> int:
    argv = argv or sys.argv[1:]
    config_path = Path(argv[0]) if argv else DEFAULT_CONFIG
    output_path = Path(argv[1]) if len(argv) > 1 else DEFAULT_OUTPUT
    generate_workbook(config_path, output_path)
    print(f"Wrote {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
